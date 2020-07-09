import urllib
import requests
from lxml import etree
from openpyxl import Workbook
from fake_useragent import UserAgent

ua = UserAgent()
wb = Workbook()

def book_spider(book_tag,page_num):
    """ 爬取书籍列表 """
    data_list = []
    for page in range(page_num):
        url = "https://book.douban.com/tag/"+urllib.parse.quote(book_tag)+"?start="+str(page*20)
        response = requests.get(url,headers={'User-Agent':ua.random})
        html = etree.HTML(response.text)
        li_list = html.xpath('//ul[@class="subject-list"]/li')
        
        for li in li_list:
            
            pic_url = li.xpath('./div[@class="pic"]//img/@src')[0]
            book_url = li.xpath('./div[@class="pic"]/a/@href')[0]
            book_name = li.xpath('./div[@class="info"]/h2/a/text()')
            book_name = ''.join(book_name).strip()
            comment_num = li.xpath('./div[@class="info"]/div[contains(@class,"star")]/span[@class="pl"]/text()')[0].strip().strip("()")
            comment_num = comment_num if len(comment_num)>0 else "暂无评论"
            book_des = li.xpath('./div[@class="info"]/p/text()')
            if len(book_des)>0:
                book_des = ''.join(book_des).strip().split('\n')
                book_des = ''.join(book_des)
            else:
                book_des = "暂无简介"
            try:
                rating = li.xpath('./div[@class="info"]/div[contains(@class,"star")]/span[@class="rating_nums"]/text()')[0].strip()
            except:
                rating = "0.0"
            book_info = li.xpath('./div[@class="info"]/div[@class="pub"]/text()')
            book_info = ''.join(book_info).strip().split('/')
            try:
                author_name = "作者/译者：" + '/'.join(book_info[0:-3])
            except:
                author_name = "作者/译者：暂无"
            try:
                pub_time = book_info[-2]
            except:
                pub_time = "无"
            try:
                price = book_info[-1]
            except:
                price = "无"
            try:
                publisher = book_info[-3]
            except:
                publisher = "无"
            data_list.append([book_name,rating,author_name,comment_num,book_des,publisher,pub_time,price,book_url,pic_url])
        print('正在爬取第{}页'.format(page))
    return data_list

def do_spider(book_tag_list,page_num):
    """ 将爬取的书籍用一个列表分类 """
    data_list = []
    for tag in book_tag_list:
        book_list = book_spider(tag,page_num)
        book_list = sorted(book_list,key=lambda x:x[1],reverse=True)
        data_list.append(book_list)
    return data_list
    
def save_to_excel(book_tag_list,book_list):
    ws = []
    save_path = "./newbook_list"
    for tag in book_tag_list:
        ws.append(wb.create_sheet(tag))
        save_path += '-'+tag
    for i in range(len(book_list)):
        ws[i].append(['书名',"评分",'作者',"评论数",'书简介','出版社','出版时间','价格','书链接','书图片链接'])
        for book in book_list[i]:
            ws[i].append(book)
        # print(data)
    wb.save(save_path+'.xlsx')
if __name__ == "__main__":
    book_tag_list = ['个人管理','时间管理','投资','文化','宗教']
    page_num = 5
    data_list = do_spider(book_tag_list,page_num)
    save_to_excel(book_tag_list,data_list)